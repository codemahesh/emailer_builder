from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: Product Import Template ──────────────────────────────────────────
ws = wb.active
ws.title = "Products"

# ── Column definitions ─────────────────────────────────────────────────────────
columns = [
    ("section_title",  "Section name to group products under (e.g. Trending, Sale). Leave blank to use 'Default'.",  25),
    ("sku",            "Unique product identifier. Either SKU or product_link must be filled.",                       20),
    ("product_link",   "Full product URL (https://...). Used for scraping name & image.",                             45),
    ("priority",       "Prominence in email. Allowed values: high, medium, low. Default: medium.",                    12),
    ("raw_price",      "Price in any format: ₹4,999 / $49.99 / 4999 INR. System normalises automatically.",          18),
    ("utm_campaign",   "Campaign tag appended to product URL as ?utm_campaign=<value>.",                              22),
    ("button_name",    "CTA button label shown in the email (e.g. Shop Now, Buy Now, Learn More).",                   18),
]

# ── Styles ─────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
REQUIRED_FILL = PatternFill("solid", fgColor="D32F2F")
OPTIONAL_FILL = PatternFill("solid", fgColor="1565C0")
AUTO_FILL     = PatternFill("solid", fgColor="2E7D32")
ALT_ROW_FILL  = PatternFill("solid", fgColor="F0F4FF")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
LABEL_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
DATA_FONT     = Font(name="Calibri", size=10)
NOTE_FONT     = Font(name="Calibri", italic=True, color="555555", size=9)

thin = Side(style="thin", color="D0D7E3")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_W  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center")

# Required / optional markers per column
REQUIRED_COLS = {"sku", "product_link"}   # one of these must be present
AUTO_COLS     = set()                      # computed by the system

# ── Row 1: column labels (Required / Optional) ─────────────────────────────────
for col_idx, (col_name, _, _col_width) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx)
    if col_name in REQUIRED_COLS:
        cell.value = "REQUIRED *"
        cell.fill  = REQUIRED_FILL
    else:
        cell.value = "OPTIONAL"
        cell.fill  = OPTIONAL_FILL
    cell.font      = LABEL_FONT
    cell.alignment = CENTER
    cell.border    = BORDER

# ── Row 2: column headers ──────────────────────────────────────────────────────
for col_idx, (col_name, _, _) in enumerate(columns, start=1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value     = col_name
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = CENTER
    cell.border    = BORDER

# ── Row 3: description row ─────────────────────────────────────────────────────
for col_idx, (_, description, _) in enumerate(columns, start=1):
    cell = ws.cell(row=3, column=col_idx)
    cell.value     = description
    cell.fill      = PatternFill("solid", fgColor="E8EDF5")
    cell.font      = NOTE_FONT
    cell.alignment = LEFT_W
    cell.border    = BORDER

ws.row_dimensions[3].height = 40

# ── Sample data rows 4–8 ──────────────────────────────────────────────────────
sample_rows = [
    ["Hero Products",   "SKU001", "https://shop.example.com/product/sneakers-v2",          "high",   "₹4,999",  "summer_launch",  "Shop Now"],
    ["Hero Products",   "SKU002", "https://shop.example.com/product/running-shoes",        "high",   "$89.99",  "summer_launch",  "Buy Now"],
    ["Trending",        "SKU003", "https://shop.example.com/product/casual-tee",           "medium", "₹999",    "may_sale",       "Explore"],
    ["Trending",        "SKU004", "https://shop.example.com/product/slim-fit-jeans",       "medium", "£45.00",  "may_sale",       ""],
    ["Sale",            "SKU005", "https://shop.example.com/product/backpack-lite?ref=em", "low",    "₹1,499",  "",               "Grab the Deal"],
]

for row_offset, row_data in enumerate(sample_rows):
    row_num = 4 + row_offset
    fill = ALT_ROW_FILL if row_offset % 2 == 0 else WHITE_FILL
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value     = value
        cell.fill      = fill
        cell.font      = DATA_FONT
        cell.alignment = LEFT
        cell.border    = BORDER

# ── Data validation: priority dropdown for column 4 (D) ───────────────────────
dv = DataValidation(
    type="list",
    formula1='"high,medium,low"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Invalid priority",
    error='Please enter: high, medium, or low',
)
ws.add_data_validation(dv)
dv.sqref = "D4:D1000"

# ── Column widths ──────────────────────────────────────────────────────────────
for col_idx, (_, _, width) in enumerate(columns, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Freeze panes below header + description rows ───────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter on header row ──────────────────────────────────────────────────
ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}2"

# ── Sheet 2: Legend / Instructions ────────────────────────────────────────────
legend = wb.create_sheet("Instructions")
legend.column_dimensions["A"].width = 22
legend.column_dimensions["B"].width = 70

instructions = [
    ("EMAIL BUILDER — PRODUCT IMPORT TEMPLATE", ""),
    ("", ""),
    ("HOW TO USE", ""),
    ("1. Fill Sheet", "Add your products starting from row 4 in the 'Products' sheet."),
    ("2. Required fields", "Each row needs at least one of: sku  OR  product_link."),
    ("3. Sections", "Group products by filling section_title. Rows with the same section_title are grouped together in the email."),
    ("4. Priority", "Controls product size/prominence in the email layout.  high → large, medium → standard, low → compact."),
    ("5. Price", "Enter price in any format — the system auto-formats it. Examples: ₹4999, $49.99, 4999 INR, EUR 12.50."),
    ("6. UTM", "Enter just the campaign name (e.g. summer_sale). The system appends ?utm_campaign=<value> to the product URL automatically."),
    ("7. Button", "Optional CTA label. If blank, the email template uses its default button text."),
    ("8. Upload", "Paste your Google Sheet URL into the campaign and click 'Full Sync' to import."),
    ("", ""),
    ("COLUMN REFERENCE", ""),
    ("section_title",  "Optional — groups products into labelled sections. Blank rows go into 'Default'."),
    ("sku",            "Stock Keeping Unit. Required if product_link is absent. Used for Fast Sync matching."),
    ("product_link",   "Full product URL. Required if sku is absent. The scraper fetches the product name and image from this URL."),
    ("priority",       "Values: high | medium | low   (default: medium)"),
    ("raw_price",      "Any price string. Auto-converted to formatted_price by the system."),
    ("utm_campaign",   "Campaign identifier appended as a UTM parameter to the product URL."),
    ("button_name",    "CTA text shown on the email button for this product."),
    ("", ""),
    ("AUTO-GENERATED (do not include in sheet)", ""),
    ("formatted_price","Calculated from raw_price. Do not add this column to the sheet."),
    ("utm_stitched",   "Built from product_link + utm_campaign. Do not add this column to the sheet."),
    ("", ""),
    ("SYNC MODES", ""),
    ("Full Sync",      "Deletes all existing products for the campaign and reimports from sheet. Triggers image scraping."),
    ("Fast Sync",      "Updates only price and UTM fields for existing products matched by SKU. Does not scrape."),
    ("", ""),
    ("SUPPORTED CURRENCIES", "USD · EUR · GBP · INR · JPY · CAD · AUD · CNY · MXN · BRL · SGD · CHF · HKD · SEK · NOK · DKK · NZD · ZAR · AED · SAR"),
]

SECTION_FILL  = PatternFill("solid", fgColor="1E3A5F")
SECTION_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
KEY_FONT      = Font(name="Calibri", bold=True, size=10)
VALUE_FONT    = Font(name="Calibri", size=10)
TITLE_FONT    = Font(name="Calibri", bold=True, color="1E3A5F", size=14)

section_keywords = {"HOW TO USE", "COLUMN REFERENCE", "SYNC MODES",
                    "AUTO-GENERATED (do not include in sheet)", "SUPPORTED CURRENCIES",
                    "EMAIL BUILDER — PRODUCT IMPORT TEMPLATE"}

for row_idx, (key, value) in enumerate(instructions, start=1):
    ka = legend.cell(row=row_idx, column=1)
    va = legend.cell(row=row_idx, column=2)
    ka.value = key
    va.value = value

    if key in section_keywords:
        ka.fill = SECTION_FILL
        va.fill = SECTION_FILL
        ka.font = SECTION_FONT
        va.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    elif key.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.")):
        ka.font = KEY_FONT
        va.font = VALUE_FONT
        ka.fill = PatternFill("solid", fgColor="EEF3FB")
        va.fill = PatternFill("solid", fgColor="EEF3FB")
    else:
        ka.font = KEY_FONT
        va.font = VALUE_FONT

    ka.alignment = LEFT_W
    va.alignment = LEFT_W
    legend.row_dimensions[row_idx].height = 22

legend.row_dimensions[1].height = 28

# ── Save ───────────────────────────────────────────────────────────────────────
out_path = r"D:\emailer_builder\Product_Import_Template.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
