"""
generate_sheet_templates.py
===========================
Generates static/sheet-template.xlsx and static/sheet-template.csv.
Run from the backend/ directory: python scripts/generate_sheet_templates.py
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ── Column definitions ────────────────────────────────────────────────────────

REQUIRED = ["sku", "product_link"]

OPTIONAL = [
    "section_title",
    "priority",
    "price",
    "utm_campaign",
    "button_name",
]

ALL_HEADERS = REQUIRED + OPTIONAL

EXAMPLE_ROWS = [
    {
        "sku": "SKU-001",
        "product_link": "https://example.com/product-1",
        "section_title": "Featured",
        "priority": "high",
        "price": "₹999",
        "utm_campaign": "summer_sale",
        "button_name": "Shop Now",
    },
    {
        "sku": "SKU-002",
        "product_link": "https://example.com/product-2",
        "section_title": "New Arrivals",
        "priority": "medium",
        "price": "₹499",
        "utm_campaign": "summer_sale",
        "button_name": "Buy Now",
    },
]

ALIASES = {
    "sku": ["sku"],
    "product_link": ["product_link", "product link", "url", "link"],
    "section_title": ["section_title", "section title", "section"],
    "priority": ["priority"],
    "price": ["price", "raw_price", "raw price"],
    "utm_campaign": ["utm_campaign", "utm campaign", "campaign"],
    "button_name": ["button_name", "button name", "button", "cta"],
}

VALID_PRIORITIES = ["high", "medium", "low"]

INSTRUCTIONS = [
    ("Column", "Required?", "Accepted aliases", "Notes"),
    ("sku", "Yes", "sku", "Unique product identifier"),
    ("product_link", "Yes", "product_link, product link, url, link", "Full URL to the product page"),
    ("section_title", "No", "section_title, section title, section", "Groups products under a heading"),
    ("priority", "No", "priority", f"One of: {', '.join(VALID_PRIORITIES)}. Defaults to medium"),
    ("price", "No", "price, raw_price, raw price", "Raw price string, e.g. ₹999 or $19.99"),
    ("utm_campaign", "No", "utm_campaign, utm campaign, campaign", "UTM campaign tag appended to product URLs"),
    ("button_name", "No", "button_name, button name, button, cta", "CTA button label, e.g. Shop Now"),
]


def _make_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()

    # ── Products sheet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Products"

    header_font = Font(bold=True, color="FFFFFF")
    req_fill = PatternFill("solid", fgColor="1E40AF")   # blue for required
    opt_fill = PatternFill("solid", fgColor="047857")   # green for optional

    for col_idx, col_name in enumerate(ALL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = req_fill if col_name in REQUIRED else opt_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, col_name in enumerate(ALL_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # Auto-width
    for col_idx, col_name in enumerate(ALL_HEADERS, start=1):
        max_len = max(len(col_name), max(len(str(r.get(col_name, ""))) for r in EXAMPLE_ROWS))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Instructions sheet ────────────────────────────────────────────────────
    ins = wb.create_sheet("Instructions")
    ins.title = "Instructions"

    hdr_fill = PatternFill("solid", fgColor="374151")
    hdr_font = Font(bold=True, color="FFFFFF")

    for col_idx, heading in enumerate(INSTRUCTIONS[0], start=1):
        cell = ins.cell(row=1, column=col_idx, value=heading)
        cell.font = hdr_font
        cell.fill = hdr_fill

    for row_idx, row in enumerate(INSTRUCTIONS[1:], start=2):
        for col_idx, value in enumerate(row, start=1):
            ins.cell(row=row_idx, column=col_idx, value=value)

    ins.column_dimensions["A"].width = 18
    ins.column_dimensions["B"].width = 12
    ins.column_dimensions["C"].width = 42
    ins.column_dimensions["D"].width = 50

    wb.save(path)
    print(f"  wrote {path}")


def _make_csv(path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ALL_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(EXAMPLE_ROWS)
    print(f"  wrote {path}")


if __name__ == "__main__":
    static = Path(__file__).parent.parent / "static"
    static.mkdir(parents=True, exist_ok=True)
    _make_xlsx(static / "sheet-template.xlsx")
    _make_csv(static / "sheet-template.csv")
    print("Done.")
